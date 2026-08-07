"""Export analysis results to a color-formatted Excel workbook."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

GREEN = "63BE7B"
YELLOW = "FFEB84"
RED = "F8696B"

# 2026-07-30: find_profitable_sizings returns every decision-point row (just
# sorted + annotated, no filtering) -- fine on the original 841k-hand dataset,
# but the 3.56M-hand dataset produces more fold-equity decision rows than
# Excel's hard per-sheet row limit (1,048,576), which crashed the report step
# with an openpyxl ValueError. The report's whole point is "which sizings are
# most profitable", so keeping only the (already sorted-descending) top slice
# is both the fix and arguably a clearer report than a multi-million-row sheet.
MAX_SIZING_ROWS = 100_000


def _write_df(ws: Worksheet, df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)


def _color_scale_column(ws: Worksheet, df: pd.DataFrame, column: str) -> None:
    if column not in df.columns:
        return
    col_idx = df.columns.get_loc(column) + 1
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    n_rows = len(df) + 1
    rule = ColorScaleRule(
        start_type="min", start_color=RED,
        mid_type="percentile", mid_value=50, mid_color=YELLOW,
        end_type="max", end_color=GREEN,
    )
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{n_rows}", rule)


def generate_report(
    output_path: str,
    player_stats_df: pd.DataFrame | None = None,
    sizing_edge_df: pd.DataFrame | None = None,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    if player_stats_df is not None and not player_stats_df.empty:
        ws = wb.create_sheet("Player Stats")
        _write_df(ws, player_stats_df)
        for col in ("vpip", "pfr", "aggression_factor", "winrate_bb_per_100"):
            _color_scale_column(ws, player_stats_df, col)

    if sizing_edge_df is not None and not sizing_edge_df.empty:
        sizing_sheet_df = sizing_edge_df.head(MAX_SIZING_ROWS)
        ws = wb.create_sheet("Profitable Sizings")
        _write_df(ws, sizing_sheet_df)
        _color_scale_column(ws, sizing_sheet_df, "edge_over_breakeven")
        _color_scale_column(ws, sizing_sheet_df, "predicted_fold_prob")

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
